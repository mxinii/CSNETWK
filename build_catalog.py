"""
Builds card_instances.json — the shared, out-of-band card catalog referenced
by RFC 0001 Section 1 ("Card costs, effects, power, toughness, and ability
text are assumed to be pre-loaded... from a shared out-of-band card catalog").

Source: mtgnp_master_card_list.xlsx
  - "Master Card List" sheet: base card stats (cost, P/T, effect text) keyed
    by Card ID Base (e.g. "mountain")
  - "Card Instances" sheet: every individual printed copy, with the exact
    card_id string used in PDUs (e.g. "mountain_001")

Output: card_instances.json — dict keyed by full instance card_id, merging
in the base card's stats. This is what server.py and client.py load, and
it's what deck_list entries must match against.

Run this once (or whenever the xlsx changes):
    python build_catalog.py
"""

import json
import re
import openpyxl

SOURCE_XLSX = "mtgnp_master_card_list.xlsx"
OUTPUT_JSON = "card_instances.json"

wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)

# --- Step 1: load base card stats, keyed by Card ID Base ---
master_ws = wb["Master Card List"]
rows = list(master_ws.iter_rows(values_only=True))
header = rows[1]  # row 0 is the title banner, row 1 is the real header
col = {name: i for i, name in enumerate(header)}

base_cards = {}
for row in rows[2:]:
    if row[col["Card ID Base"]] is None:
        continue
    base_id = row[col["Card ID Base"]]
    base_cards[base_id] = {
        "card_name": row[col["Card Name"]],
        "card_type": row[col["Card Type"]],
        "subtype": row[col["Subtype"]],
        "color": row[col["Color"]],
        "cmc": row[col["CMC"]],
        "cost": {
            "W": row[col["W"]],
            "U": row[col["U"]],
            "B": row[col["B"]],
            "R": row[col["R"]],
            "G": row[col["G"]],
            "Generic": row[col["Generic"]],
        },
        "power": row[col["Power"]],
        "toughness": row[col["Toughness"]],
        "effect": row[col["Simplified Effect"]],
    }

# --- Step 2: load every instance, merge in base stats ---
instances_ws = wb["Card Instances"]
rows = list(instances_ws.iter_rows(values_only=True))
header = rows[1]
col = {name: i for i, name in enumerate(header)}

card_instances = {}
for row in rows[2:]:
    instance_id = row[col["card_id (protocol reference)"]]
    if instance_id is None:
        continue

    # Derive the base ID by stripping the trailing "_NNN" copy number
    base_id = re.sub(r"_\d+$", "", instance_id)
    base_info = base_cards.get(base_id)
    if base_info is None:
        raise ValueError(f"Instance '{instance_id}' has no matching base card '{base_id}'")

    entry = dict(base_info)  # copy so each instance has its own dict
    entry["card_id"] = instance_id
    entry["copy_number"] = row[col["Copy #"]]
    card_instances[instance_id] = entry

with open(OUTPUT_JSON, "w") as f:
    json.dump(card_instances, f, indent=2)

print(f"Wrote {len(card_instances)} card instances to {OUTPUT_JSON}")
